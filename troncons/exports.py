import io
from datetime import date

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_pdf(troncon, resultat):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', fontSize=16, spaceAfter=6, alignment=TA_CENTER,
                                  textColor=colors.HexColor('#1e3a5f'), fontName='Helvetica-Bold')
    h2_style = ParagraphStyle('H2', fontSize=12, spaceBefore=12, spaceAfter=4,
                               textColor=colors.HexColor('#2563EB'), fontName='Helvetica-Bold')
    body_style = styles['Normal']

    RISQUE_COLOR = {
        'Critique': colors.HexColor('#DC2626'),
        'Élevé': colors.HexColor('#D97706'),
        'Modéré': colors.HexColor('#16A34A'),
        'Modéré (stable)': colors.HexColor('#16A34A'),
    }

    story = []

    # En-tête
    story.append(Paragraph("VIZIR — Rapport de Dégradation de Chaussée", title_style))
    story.append(Paragraph(f"Tronçon : {troncon.nom} — {troncon.route or ''}", h2_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#2563EB')))
    story.append(Spacer(1, 0.3*cm))

    # Informations générales
    story.append(Paragraph("1. Identification du tronçon", h2_style))
    data_info = [
        ['Route', troncon.route or '—'],
        ['PK début / PK fin', f"{troncon.pk_debut or '—'} / {troncon.pk_fin or '—'}"],
        ['Longueur', f"{troncon.longueur_m or '—'} m"],
        ['Localisation', troncon.localisation or '—'],
        ['Date d\'inspection', str(resultat.parametres.date_inspection) if resultat.parametres else '—'],
        ['Date de calcul', str(resultat.date_calcul.strftime('%d/%m/%Y %H:%M'))],
    ]
    t = Table(data_info, colWidths=[6*cm, 11*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EFF6FF')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F0F9FF')]),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.4*cm))

    # Paramètres d'entrée
    story.append(Paragraph("2. Paramètres d'entrée", h2_style))
    p = resultat.parametres
    A_val = p.get_A() if p else '—'
    CL_val = p.get_CL() if p else '—'
    M_val = p.get_M() if p else '—'
    data_params = [
        ['Paramètre', 'Valeur', 'Mode de saisie'],
        ['Indice VIZIR initial (ID₀)', str(p.id0_vizir if p else '—'), '—'],
        ['Agressivité du trafic (A)', f"{A_val:.3f}" if isinstance(A_val, float) else '—', p.mode_A if p else '—'],
        ['Coefficient climatique (CL)', f"{CL_val:.3f}" if isinstance(CL_val, float) else '—', p.mode_CL if p else '—'],
        ['Paramètre matériau (M)', f"{M_val:.3f}" if isinstance(M_val, float) else '—', p.mode_M if p else '—'],
    ]
    t2 = Table(data_params, colWidths=[7*cm, 5*cm, 5*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F9FF')]),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t2)
    story.append(Spacer(1, 0.4*cm))

    # Résultats
    story.append(Paragraph("3. Résultats du modèle VIZIR", h2_style))
    risque_color = RISQUE_COLOR.get(resultat.niveau_risque, colors.black)
    data_res = [
        ['Indicateur', 'Valeur'],
        ['Rythme de dégradation R', f"{resultat.R:.3f} pts/an"],
        ['Contribution Trafic (αA)', f"{resultat.contrib_trafic:.3f} ({resultat.contrib_trafic_pct:.1f}%)"],
        ['Contribution Climat (βCL)', f"{resultat.contrib_climat:.3f} ({resultat.contrib_climat_pct:.1f}%)"],
        ['Contribution Matériau (−γM)', f"{resultat.contrib_materiau:.3f} ({resultat.contrib_materiau_pct:.1f}%)"],
        ['Facteur dominant', resultat.facteur_dominant],
        ['Niveau de risque', resultat.niveau_risque],
        ['Temps avant ID=7 (T_critique)', f"{resultat.t_critique_ans:.1f} ans" if resultat.t_critique_ans is not None else "Stable"],
    ]
    t3 = Table(data_res, colWidths=[9*cm, 8*cm])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F9FF')]),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('TEXTCOLOR', (1, 6), (1, 6), risque_color),
        ('FONTNAME', (1, 6), (1, 6), 'Helvetica-Bold'),
    ]))
    story.append(t3)
    story.append(Spacer(1, 0.4*cm))

    # Projections
    story.append(Paragraph("4. Projection temporelle ID(t)", h2_style))
    projections = resultat.get_projections_display()
    proj_data = [['Horizon (années)', 'ID projeté', 'Statut']]
    for h, v in projections:
        statut = "⚠ Critique" if v >= 7 else ("Dégradé" if v >= 5 else "Acceptable")
        proj_data.append([f"{h} ans", f"{v:.2f}", statut])
    t4 = Table(proj_data, colWidths=[6*cm, 6*cm, 5*cm])
    t4.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F9FF')]),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t4)
    story.append(Spacer(1, 0.4*cm))

    # Recommandations
    story.append(Paragraph("5. Recommandations de gestion", h2_style))
    story.append(Paragraph(f"<b>Facteur dominant :</b> {resultat.facteur_dominant}", body_style))
    story.append(Spacer(1, 0.2*cm))
    for reco in resultat.recommandations.split('\n'):
        if reco.strip():
            story.append(Paragraph(f"• {reco.strip()}", body_style))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph(f"Rapport généré le {date.today().strftime('%d/%m/%Y')} — Modèle VIZIR",
                            ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(story)
    buffer.seek(0)
    return buffer


def export_excel(donnees):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparatif VIZIR"

    # Styles
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    RISQUE_FILLS = {
        'Critique': PatternFill("solid", fgColor="DC2626"),
        'Élevé': PatternFill("solid", fgColor="D97706"),
        'Modéré': PatternFill("solid", fgColor="16A34A"),
        'Modéré (stable)': PatternFill("solid", fgColor="16A34A"),
    }

    # Titre (fusionné sur 13 colonnes)
    ws.merge_cells('A1:M1')
    ws['A1'] = "VIZIR — Tableau Comparatif des Tronçons"
    ws['A1'].font = Font(bold=True, size=14, color="1e3a5f")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 25

    # En-têtes ligne 2
    headers = [
        'Tronçon', 'Route', 'ID₀', 'A', 'CL', 'M',
        'R (pts/an)', 'ID à 5 ans', 'ID à 10 ans',
        'Facteur dominant', 'Niveau de risque', 'T_critique (ans)', 'Action recommandée'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    ws.row_dimensions[2].height = 20

    # Données
    for row_idx, d in enumerate(donnees, 3):
        t = d['troncon']
        r = d['resultat']
        p = r.parametres

        projs = {float(k): v for k, v in r.projections_json.items()}
        id5 = projs.get(5.0, '—')
        id10 = projs.get(10.0, '—')

        values = [
            t.nom,
            t.route or '—',
            p.id0_vizir if p else '—',
            round(p.get_A(), 3) if p and p.get_A() else '—',
            round(p.get_CL(), 3) if p and p.get_CL() else '—',
            round(p.get_M(), 3) if p and p.get_M() else '—',
            round(r.R, 3),
            round(id5, 2) if isinstance(id5, float) else id5,
            round(id10, 2) if isinstance(id10, float) else id10,
            r.facteur_dominant,
            r.niveau_risque,
            round(r.t_critique_ans, 1) if r.t_critique_ans is not None else 'Stable',
            r.recommandations.split('\n')[0] if r.recommandations else '—',
        ]

        alt_fill = PatternFill("solid", fgColor="EFF6FF") if row_idx % 2 == 0 else None

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            if alt_fill:
                cell.fill = alt_fill
            if col == 11:
                fill = RISQUE_FILLS.get(val)
                if fill:
                    cell.fill = fill
                    cell.font = Font(color="FFFFFF", bold=True, size=9)

        ws.row_dimensions[row_idx].height = 18

    # Largeurs colonnes — get_column_letter évite le bug MergedCell
    col_widths = [12, 20, 7, 8, 8, 8, 12, 12, 12, 18, 15, 14, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer